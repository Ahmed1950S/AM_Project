import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from scipy.optimize import minimize
import warnings

warnings.filterwarnings("ignore")

# ─── Configuration ────────────────────────────────────────────────────────────
REGION = "EUR"
START_YEAR = 2013  # first allocation decision
END_YEAR = 2024  # last allocation decision
ESTIM_YEARS = 10  # rolling window length
MIN_OBS = 36  # min valid monthly returns in window
STALE_THR = 0.30  # max zero-return fraction
LOW_FLOOR = 0.50  # RI values below this → NaN
LW_SHRINK_FLOOR = 0.01  # minimum Ledoit-Wolf shrinkage intensity
DATA_PATH = "Data_2026/"
TEMPLATE_PATH = "Data_2026/Template_for_Part_I-SAAM.xlsx"
OUT = "Output_2026/"

print("=" * 65)
print("SAAM Part I — Minimum Variance Portfolio (EUR region)")
print("=" * 65)

# =============================================================================
# 1. LOAD RAW DATA
# =============================================================================
print("\n[1] Loading raw data ...")

static = pd.read_excel(DATA_PATH + "Static_2025.xlsx")
ri_m_raw = pd.read_excel(DATA_PATH + "DS_RI_T_USD_M_2025.xlsx", sheet_name="RI")
mv_m_raw = pd.read_excel(DATA_PATH + "DS_MV_T_USD_M_2025.xlsx", sheet_name="MV")
ri_y_raw = pd.read_excel(DATA_PATH + "DS_RI_T_USD_Y_2025.xlsx", sheet_name="RI")
mv_y_raw = pd.read_excel(DATA_PATH + "DS_MV_T_USD_Y_2025.xlsx", sheet_name="MV")
co2_s1_raw = pd.read_excel(DATA_PATH + "DS_CO2_SCOPE_1_Y_2025.xlsx", sheet_name="Scope1")
co2_s2_raw = pd.read_excel(DATA_PATH + "DS_CO2_SCOPE_2_Y_2025.xlsx", sheet_name="Scope2")
rev_raw = pd.read_excel(DATA_PATH + "DS_REV_Y_2025.xlsx", sheet_name="REV")
rf_raw = pd.read_excel(DATA_PATH + "Risk_Free_Rate_2025.xlsx",
                       sheet_name="F-F_Research_Data_Factors")

print(f"   Static: {static.shape}, RI monthly: {ri_m_raw.shape}")

# =============================================================================
# 2. CLEAN DATASTREAM FORMAT
# =============================================================================
print("\n[2] Cleaning Datastream format ...")


def clean_ds(df, valid_isins):
    """Drop error rows, filter to valid ISINs, set ISIN as index, coerce to numeric."""
    df = df.dropna(subset=["ISIN"]).copy()
    df = df[~df["ISIN"].astype(str).str.contains("ER:", na=False)]
    df = df[df["ISIN"].isin(valid_isins)]
    names = df.set_index("ISIN")["NAME"]
    data_cols = [c for c in df.columns if c not in ["NAME", "ISIN"] and c is not None]
    data = df.set_index("ISIN")[data_cols].apply(pd.to_numeric, errors="coerce")
    # Normalise column types: datetime for monthly, keep int for yearly
    new_cols = []
    for c in data.columns:
        if hasattr(c, "year") and not isinstance(c, (int, np.integer)):
            new_cols.append(pd.Timestamp(c))
        else:
            new_cols.append(c)
    data.columns = new_cols
    return data, names


eur_isins = set(static.loc[static["Region"] == REGION, "ISIN"])
print(f"   EUR ISINs in static: {len(eur_isins)}")

ri_m, firm_names = clean_ds(ri_m_raw, eur_isins)
mv_m, _ = clean_ds(mv_m_raw, eur_isins)
ri_y, _ = clean_ds(ri_y_raw, eur_isins)
mv_y, _ = clean_ds(mv_y_raw, eur_isins)
co2_s1, _ = clean_ds(co2_s1_raw, eur_isins)
co2_s2, _ = clean_ds(co2_s2_raw, eur_isins)
rev, _ = clean_ds(rev_raw, eur_isins)

isin_name = firm_names.to_dict()
print(f"   EUR firms loaded: {ri_m.shape[0]}")

# =============================================================================
# 3. BUILD DATE LISTS
# =============================================================================

monthly_all = sorted([c for c in ri_m.columns if isinstance(c, pd.Timestamp)
                      and pd.Timestamp("2000-01-01") <= c <= pd.Timestamp("2025-12-31")])
annual_all = sorted([c for c in ri_y.columns if isinstance(c, (int, np.integer))])

# Restrict all frames to valid date ranges
ri_m = ri_m[monthly_all]
mv_m = mv_m[[c for c in monthly_all if c in mv_m.columns]]
ri_y = ri_y[[c for c in annual_all if c in ri_y.columns]]
mv_y = mv_y[[c for c in annual_all if c in mv_y.columns]]
co2_s1 = co2_s1[[c for c in annual_all if c in co2_s1.columns]]
co2_s2 = co2_s2[[c for c in annual_all if c in co2_s2.columns]]
rev = rev[[c for c in annual_all if c in rev.columns]]


def months_of(Y):
    """Return list of monthly Timestamps for year Y."""
    return [c for c in monthly_all if c.year == Y]


def estim_window(Y):
    """Return list of monthly Timestamps in the 10-year estimation window ending Dec Y."""
    return [c for c in monthly_all
            if (Y - ESTIM_YEARS + 1, 1) <= (c.year, c.month) <= (Y, 12)]


# Verify
ew2013 = estim_window(2013)
print(f"   Estimation window Dec 2013: {len(ew2013)} months "
      f"({ew2013[0].strftime('%Y-%m')} to {ew2013[-1].strftime('%Y-%m')})")
assert len(ew2013) == 120, f"Expected 120, got {len(ew2013)}"

# =============================================================================
# 4. RISK-FREE RATE
# =============================================================================
print("\n[3] Processing risk-free rate ...")

rf_raw.columns = ["YYYYMM", "RF_pct"]
rf_raw = rf_raw.dropna(subset=["YYYYMM"])
rf_raw["date"] = pd.to_datetime(rf_raw["YYYYMM"].astype(int).astype(str), format="%Y%m")
rf_raw["date"] = rf_raw["date"] + pd.offsets.MonthEnd(0)

# RF_pct is the MONTHLY rate expressed in percent (Fama-French convention).
# Just convert percent → decimal. Annualisation happens later in compute_perf
# via rf_ann = 12 × mean(rf_monthly), to match how portfolio returns are annualised.
rf_mon = rf_raw.set_index("date")["RF_pct"] / 100
rf_mon = rf_mon.squeeze()
rf_mon.name = "RF"

print(f"   RF range: {rf_mon.index.min().strftime('%Y-%m')} to "
      f"{rf_mon.index.max().strftime('%Y-%m')}")
print(f"   RF sample 2014-01: {rf_mon.loc['2014-01'].values[0]:.6f} (monthly)")

# --- Sanity check: annualised avg rf in our sample window should be in a
# plausible T-bill range (~0.5%–5% over 2014–2025). If not, the conversion
# above is wrong. This catches the common mistake of compounding monthly
# rates as if they were annual.
_rf_ann_check = rf_mon.loc["2014-01-01":"2025-12-31"].mean() * 12
assert 0.005 < _rf_ann_check < 0.05, (
    f"Annualised avg rf = {_rf_ann_check:.4%} — outside plausible range. "
    f"Check whether RF_pct is monthly (correct) or annual (needs ^(1/12))."
)
print(f"   Sanity: annualised avg rf over 2014-2025 = {_rf_ann_check*100:.4f}%")

# =============================================================================
# 5. PRICE CLEANING
# =============================================================================
print("\n[4] Cleaning prices ...")

# 5a. Low price filter
n_low = ((ri_m > 0) & (ri_m < LOW_FLOOR)).sum().sum()
ri_m[ri_m < LOW_FLOOR] = np.nan
print(f"   Prices < {LOW_FLOOR} set to NaN: {n_low}")


# 5b. Forward-fill internal missing prices
def forward_fill_internal(row):
    fv = row.first_valid_index()
    lv = row.last_valid_index()
    if fv is None or lv is None:
        return row
    cols = list(row.index)
    start = cols.index(fv)
    end = cols.index(lv)
    row.iloc[start:end + 1] = row.iloc[start:end + 1].ffill()
    return row


ri_before = ri_m.isna().sum().sum()
ri_m = ri_m.apply(forward_fill_internal, axis=1)
n_filled = ri_before - ri_m.isna().sum().sum()
print(f"   Forward-filled internal gaps: {n_filled} observations")

# 5c. Detect delisted firms (last valid price before end of sample)
cutoff = pd.Timestamp("2025-12-31")
last_valid = {}
for isin in ri_m.index:
    lv = ri_m.loc[isin].last_valid_index()
    if lv is not None and isinstance(lv, pd.Timestamp) and lv < cutoff:
        last_valid[isin] = lv
print(f"   Firms with early last price (potential delistings): {len(last_valid)}")

# 5d. Compute returns + apply -100% at delisting
ret_m = ri_m.pct_change(axis=1)

for isin, ddate in last_valid.items():
    if ddate not in monthly_all:
        continue
    idx = monthly_all.index(ddate)
    if idx + 1 < len(monthly_all):
        ret_m.at[isin, monthly_all[idx + 1]] = -1.0
        # Set all returns after -100% to NaN
        for k in range(idx + 2, len(monthly_all)):
            if monthly_all[k] in ret_m.columns:
                ret_m.at[isin, monthly_all[k]] = np.nan

# Drop the first column (no return for the first price date)
ret_m = ret_m.iloc[:, 1:]

n_delist = (ret_m == -1.0).sum().sum()
print(f"   -100% delisting returns applied: {n_delist}")

# 5e. Forward-fill annual CO2 and revenue (per project: use previous year if missing)
co2_s1 = co2_s1.ffill(axis=1)
co2_s2 = co2_s2.ffill(axis=1)
rev = rev.ffill(axis=1)

# =============================================================================
# 6. INVESTMENT SET (UNIVERSE) CONSTRUCTION
# =============================================================================
print("\n[5] Building investment sets ...")


def get_universe(Y):
    """
    Build investment set for allocation at end of year Y.
    Filters:
      1. Valid price at end of year Y (monthly RI)
      2. Sufficient return observations in 10-year window
      3. No stale prices (zero-return fraction < threshold)
      4. CO2 Scope 1 + Scope 2 both available at end of year Y
    """
    # Dec Y price column
    dec_cols = [c for c in monthly_all if c.year == Y and c.month == 12]
    if not dec_cols:
        return []
    dec_Y = dec_cols[0]

    # Estimation window returns
    win = estim_window(Y)
    win_ret = [c for c in win if c in ret_m.columns]
    R_win = ret_m.reindex(columns=win_ret)

    out = []
    for isin in ri_m.index:
        # Filter 1: valid price at end of year Y
        if pd.isna(ri_m.at[isin, dec_Y]):
            continue

        # Filter 2: sufficient returns
        if isin not in R_win.index:
            continue
        row = R_win.loc[isin]
        n_valid = row.notna().sum()
        if n_valid < MIN_OBS:
            continue

        # Filter 3: stale prices (Fixed to precisely match data_exploration.py)
        n_zero = ((row == 0) | (row.abs() < 1e-10)).sum()
        if (n_zero / n_valid) > STALE_THR:
            continue

        # Filter 4: CO2 Scope 1 AND Scope 2 available
        has_s1 = Y in co2_s1.columns and pd.notna(co2_s1.at[isin, Y]) if isin in co2_s1.index else False
        has_s2 = Y in co2_s2.columns and pd.notna(co2_s2.at[isin, Y]) if isin in co2_s2.index else False
        if not (has_s1 and has_s2):
            continue

        out.append(isin)

    return out


universe = {}
for Y in range(START_YEAR, END_YEAR + 1):
    universe[Y] = get_universe(Y)
    print(f"   {Y}: {len(universe[Y]):3d} firms")


# =============================================================================
# 7. COVARIANCE ESTIMATION
# =============================================================================

def estimate_cov(isins, win_cols):
    """
    Estimate expected returns and covariance matrix using the lecture method,
    with Ledoit-Wolf shrinkage to constant-correlation target.

    Step 1 — Pairwise-complete covariance (Lecture 5, slides 20-21):
      - Var(Ri) computed over firm i's available sample (ML: divide by n)
      - Corr(Ri,Rj) computed over the common sample of i and j
      - Cov(Ri,Rj) = Corr(Ri,Rj) * sqrt(Var(Ri) * Var(Rj))

    Step 2 — Ledoit-Wolf shrinkage (Ledoit & Wolf, 2004):
      Σ_shrunk = δ·F + (1-δ)·S
      where F is the constant-correlation target and δ is the optimal
      shrinkage intensity estimated analytically.
      This guarantees PSD and reduces estimation error in finite samples.

    Reference: Ledoit, O. & Wolf, M. (2004), "A well-conditioned estimator
    for large-dimensional covariance matrices", Journal of Multivariate
    Analysis, 88(2), 365-411.
    """
    win_in = [c for c in win_cols if c in ret_m.columns]
    R = ret_m.loc[isins, win_in]  # (N, T) with NaN
    N = len(isins)
    T = len(win_in)

    # --- Expected returns: per-firm mean over available obs ---
    mu = R.mean(axis=1).values  # pandas mean skips NaN

    # --- Variance for each firm (ML: divide by n) ---
    R_vals = R.values  # (N, T) numpy array
    not_nan = ~np.isnan(R_vals)  # (N, T) boolean

    # Per-firm mean (only over available obs)
    mu_full = np.nanmean(R_vals, axis=1)  # (N,)

    # Per-firm variance (ML)
    R_demeaned = R_vals - mu_full[:, None]  # (N, T), NaN propagates
    R_demeaned_zero = np.where(not_nan, R_demeaned, 0.0)  # NaN → 0 for sum
    n_obs = not_nan.sum(axis=1)  # (N,) count per firm
    var = np.where(n_obs > 1,
                   np.sum(R_demeaned_zero ** 2, axis=1) / n_obs,
                   0.0)  # (N,)

    # --- Pairwise correlation (vectorized) ---
    R_zero = np.where(not_nan, R_vals, 0.0)  # (N, T)
    not_nan_f = not_nan.astype(np.float64)  # (N, T)

    # Count of common observations for each pair
    count_ij = not_nan_f @ not_nan_f.T  # (N, N)

    # Sum of R_i over common obs with j
    sum_i_ij = R_zero @ not_nan_f.T  # (N, N)
    sum_j_ij = not_nan_f @ R_zero.T  # (N, N)

    # Pairwise means over common sample
    safe_count = np.maximum(count_ij, 1)
    mean_i_ij = sum_i_ij / safe_count  # (N, N)
    mean_j_ij = sum_j_ij / safe_count  # (N, N)

    # Pairwise covariance (ML)
    cross_ij = R_zero @ R_zero.T  # (N, N): [i,j] = sum of R_i*R_j over common obs
    cov_ij = (cross_ij / safe_count) - mean_i_ij * mean_j_ij  # (N, N) pairwise cov

    # Pairwise standard deviations over common samples
    sum_sq_i_ij = (R_zero ** 2) @ not_nan_f.T  # [i,j] = sum of R_i^2 over common
    var_i_ij = sum_sq_i_ij / safe_count - mean_i_ij ** 2
    var_i_ij = np.maximum(var_i_ij, 0)  # numerical safety

    sum_sq_j_ij = not_nan_f @ (R_zero ** 2).T
    var_j_ij = sum_sq_j_ij / safe_count - mean_j_ij ** 2
    var_j_ij = np.maximum(var_j_ij, 0)

    # Correlation
    denom = np.sqrt(var_i_ij * var_j_ij)
    corr_ij = np.where(denom > 1e-20, cov_ij / denom, 0.0)

    # Reconstruct covariance: Cov(i,j) = rho(i,j) * sigma_i * sigma_j
    std_own = np.sqrt(var)  # (N,)
    Sig = corr_ij * np.outer(std_own, std_own)

    # Set diagonal to own variance
    np.fill_diagonal(Sig, var)

    # Enforce symmetry (numerical)
    Sig = (Sig + Sig.T) / 2

    # =========================================================================
    # Ledoit-Wolf shrinkage to constant-correlation target
    # =========================================================================
    # Step 2a: Build the structured target F (constant correlation, own variances)
    # Extract the correlation matrix from Sig
    std_diag = np.sqrt(np.diag(Sig))
    std_diag_safe = np.where(std_diag > 1e-20, std_diag, 1e-20)
    corr_mat = Sig / np.outer(std_diag_safe, std_diag_safe)
    np.fill_diagonal(corr_mat, 1.0)
    # Clip correlations to [-1, 1] for numerical safety
    corr_mat = np.clip(corr_mat, -1.0, 1.0)

    # Average off-diagonal correlation
    rho_bar = (corr_mat.sum() - N) / (N * (N - 1))

    # Constant-correlation target F: F_ij = rho_bar * sigma_i * sigma_j for i≠j
    F = rho_bar * np.outer(std_diag, std_diag)
    np.fill_diagonal(F, var)  # F_ii = Var(Ri)

    # Step 2b: Estimate the optimal shrinkage intensity δ
    # Following Ledoit & Wolf (2004), we estimate δ = κ/T where κ depends on
    # π (sum of asymptotic variances of entries of S),
    # γ (distance between S and F), and ρ (a cross term).
    # For the constant-correlation target, we use the simplified estimator
    # from "Honey, I Shrunk the Sample Covariance Matrix" (Ledoit & Wolf, 2003).
    #
    # Approximate δ using the formula: δ* ≈ (sum of Var(s_ij)) / ||S - F||²
    # computed over the available return data.

    # Compute π̂ (sum of squared deviations of sample cov entries from their means)
    # We use the returns matrix with NaN filled to 0 and masked,
    # then compute the per-element variance of the outer products.
    # Simplified Ledoit-Wolf: use a consistent estimator via the returns.

    # For computational tractability with missing data, we use the Oracle
    # Approximating Shrinkage (OAS) formula which is simpler and robust:
    #   δ_OAS = ( (1-2/N)*tr(Σ²) + tr²(Σ) ) / ( (T+1-2/N)*(tr(Σ²) - tr²(Σ)/N) )
    # Reference: Chen, Y., Wiesel, A., Eldar, Y., & Hero, A. (2010),
    # "Shrinkage Algorithms for MMSE Covariance Estimation",
    # IEEE Transactions on Signal Processing, 58(10), 5016-5029.

    tr_Sig = np.trace(Sig)
    tr_Sig2 = np.trace(Sig @ Sig)
    # Use the effective sample size (median pairwise count) as T_eff
    T_eff = np.median(count_ij[np.triu_indices(N, k=1)])
    T_eff = max(T_eff, 2)  # safety

    numerator = (1.0 - 2.0 / N) * tr_Sig2 + tr_Sig ** 2
    denominator = (T_eff + 1.0 - 2.0 / N) * (tr_Sig2 - tr_Sig ** 2 / N)

    if abs(denominator) < 1e-20:
        delta = 0.5  # fallback
    else:
        delta = max(min(numerator / denominator, 1.0), LW_SHRINK_FLOOR)

    # Step 2c: Apply shrinkage
    Sig_shrunk = delta * F + (1.0 - delta) * Sig

    # Final PSD enforcement via spectral clipping (safety net)
    eigvals, eigvecs = np.linalg.eigh(Sig_shrunk)
    if eigvals[0] < 1e-10:
        eigvals = np.maximum(eigvals, 1e-10)
        Sig_shrunk = eigvecs @ np.diag(eigvals) @ eigvecs.T
        Sig_shrunk = (Sig_shrunk + Sig_shrunk.T) / 2  # enforce symmetry

    return mu, Sig_shrunk


# =============================================================================
# 8. MIN-VARIANCE OPTIMISATION
# =============================================================================
print("\n[6] Rolling min-variance optimisation ...")

_drifted_w = {}  # year → dict of {isin: drifted_weight} at end of year


def min_var_weights(Sigma, isins, Y):
    """
    Solve: min α'Σα  s.t. α'e = 1, α >= 0

    Warm start: use the drifted portfolio weights at end of year Y
    (i.e., the actual portfolio composition after 12 months of price drift).
    This is what the investor is currently holding before rebalancing.
    For the first year, use equal weights 1/N.
    """
    N = Sigma.shape[0]

    # Warm start from drifted weights (actual holdings before rebalancing)
    if Y in _drifted_w:
        prev = _drifted_w[Y]
        w0 = np.array([prev.get(i, 0.0) for i in isins])
        w0 = w0 / w0.sum() if w0.sum() > 0 else np.ones(N) / N
    else:
        w0 = np.ones(N) / N

    res = minimize(
        fun=lambda w: float(w @ Sigma @ w),
        x0=w0,
        jac=lambda w: 2.0 * (Sigma @ w),
        method="SLSQP",
        bounds=[(0.0, None)] * N,
        constraints={"type": "eq", "fun": lambda w: w.sum() - 1.0},
        options={"ftol": 1e-10, "maxiter": 1000},
    )

    if not res.success:
        print(f"   WARNING: optimizer did not converge for Y={Y}: {res.message}")

    return res.x


# Rolling loop
mv_w_dict = {}  # year → pd.Series of optimal weights
mv_ret = {}  # year+1 → pd.Series of monthly portfolio returns


def fill_oos_returns(eligible, next_months):
    """
    Fill missing OOS returns with proper delisting detection.

    Logic (per project instructions, Section 1 — Data Cleaning):
      - If a firm had a valid price last month but price is now missing,
        treat as delisting → return = -100%, all subsequent months = NaN.
      - If the firm had no valid price last month either (leading NaN),
        treat as 0% return (firm not yet active in this sub-period).
      - Remaining NaN after delisting logic → 0% (no price change).

    Reference: Shumway, T. (1997), "The Delisting Bias in CRSP Data",
    Journal of Finance, 52(1), 327-340.
    """
    R_oos = ret_m.loc[eligible].reindex(columns=next_months).copy()

    for isin in eligible:
        delisted = False
        for k, t in enumerate(next_months):
            if delisted:
                R_oos.at[isin, t] = np.nan
                continue

            if pd.isna(R_oos.at[isin, t]):
                # Check if previous month had a valid RI price
                t_idx = monthly_all.index(t)
                prev_t = monthly_all[t_idx - 1] if t_idx > 0 else None

                had_price_prev = (prev_t is not None
                                  and prev_t in ri_m.columns
                                  and isin in ri_m.index
                                  and pd.notna(ri_m.at[isin, prev_t]))

                if had_price_prev:
                    # Price disappeared → delisting: realised return = -100%
                    R_oos.at[isin, t] = -1.0
                    delisted = True
                else:
                    # No previous price either → treat as 0% (inactive)
                    R_oos.at[isin, t] = 0.0

    # Any remaining NaN (e.g., leading gaps) → 0%
    R_oos = R_oos.fillna(0.0)
    return R_oos

for Y in range(START_YEAR, END_YEAR + 1):
    eligible = universe[Y]
    N = len(eligible)
    if N == 0:
        print(f"   Y={Y}: 0 firms → skip")
        continue

    # Estimate moments
    mu, Sig = estimate_cov(eligible, estim_window(Y))

    # Optimise
    w = min_var_weights(Sig, eligible, Y)
    mv_w_dict[Y] = pd.Series(w, index=eligible)

    # Ex-ante annualised volatility
    ea_vol = np.sqrt(float(w @ Sig @ w) * 12) * 100

    # Compute ex-post returns for year Y+1, with weight drift
    next_months = months_of(Y + 1)
    R_next = fill_oos_returns(eligible, next_months)
    ww = w.copy()
    port_ret = []
    for t in next_months:
        r_t = R_next[t].values
        rp_t = float(ww @ r_t)
        port_ret.append(rp_t)
        # Weight drift: α_{i,t+k-1} = α_{i,t+k-2} * (1+R_{i,t+k-1}) / (1+R_{p,t+k-1})
        ww = ww * (1.0 + r_t) / max(1.0 + rp_t, 1e-12)
    mv_ret[Y + 1] = pd.Series(port_ret, index=next_months)

    # Store drifted weights at end of year Y+1 for next year's warm start
    _drifted_w[Y + 1] = dict(zip(eligible, ww))

    n_nonzero = (w > 1e-6).sum()
    print(f"   Y={Y}: {N:3d} firms, non-zero wgts: {n_nonzero:3d}, "
          f"ex-ante ann.σ: {ea_vol:.2f}%")

# Concatenate into a single return series
rp_mv = pd.concat(mv_ret).droplevel(0).sort_index()
rp_mv.index = pd.DatetimeIndex(rp_mv.index)

# =============================================================================
# 9. VALUE-WEIGHTED BENCHMARK
# =============================================================================
print("\n[7] Value-weighted benchmark ...")

vw_ret = {}
for Y in range(START_YEAR, END_YEAR + 1):
    eligible = universe[Y]
    next_months = months_of(Y + 1)
    # Pre-compute delisting-aware returns for this OOS year
    R_oos_vw = fill_oos_returns(eligible, next_months)
    port = []

    for t in next_months:
        # Use previous month's market cap as weights
        idx = monthly_all.index(t) if t in monthly_all else None
        if idx is None or idx == 0:
            port.append(np.nan)
            continue
        prev_t = monthly_all[idx - 1]

        cap = (mv_m.loc[eligible, prev_t].fillna(0)
               if prev_t in mv_m.columns else pd.Series(0.0, index=eligible))
        tot = cap.sum()
        if tot <= 0:
            port.append(0.0)
            continue

        r_t = R_oos_vw[t].values if t in R_oos_vw.columns else np.zeros(len(eligible))
        port.append(float((cap / tot).values @ r_t))

    vw_ret[Y + 1] = pd.Series(port, index=next_months)

rp_vw = pd.concat(vw_ret).droplevel(0).sort_index()
rp_vw.index = pd.DatetimeIndex(rp_vw.index)

# =============================================================================
# 10. PERFORMANCE STATISTICS
# =============================================================================
print("\n[8] Performance statistics ...")


def compute_perf(rp, rf_s, label):
    """Compute standard performance metrics.

    Annualised return:
      - Arithmetic: 12 × mean(R_monthly)                    [Lecture 5, slide 12]
      - Geometric:  (1 + R_cum)^(12/T) − 1                  [Lecture 5, slide 13]
    Sharpe ratio uses the arithmetic return for consistency with
    the annualisation SR^(y) = sqrt(12) × SR^(m)             [Lecture 5, slide 12]
    """
    rp = rp.dropna()
    rf = rf_s.reindex(rp.index).ffill().fillna(0)
    T = len(rp)

    # Arithmetic annualised return: R̄_p^(y) = 12 × R̄_p^(m)
    mu_arith = 12 * rp.mean()
    # Geometric annualised return: (1 + R_cum)^(12/T) − 1
    mu_geom = (1 + rp).prod() ** (12 / T) - 1
    # Annualised volatility: σ_p^(y) = √12 × σ_p^(m)
    sig_ann = rp.std() * np.sqrt(12)
    # Arithmetic annualised risk-free rate (same convention)
    rf_ann = 12 * rf.mean()
    # Sharpe ratio (arithmetic, consistent with slide 12)
    SR = (mu_arith - rf_ann) / sig_ann
    # Drawdown
    cum = (1 + rp).cumprod()
    mdd = ((cum - cum.cummax()) / cum.cummax()).min()

    return {
        "Portfolio": label,
        "Ann. Return Arith. (%)": round(mu_arith * 100, 2),
        "Ann. Return Geom. (%)": round(mu_geom * 100, 2),
        "Ann. Vol (%)": round(sig_ann * 100, 2),
        "Sharpe": round(SR, 3),
        "Min Mo. (%)": round(rp.min() * 100, 2),
        "Max Mo. (%)": round(rp.max() * 100, 2),
        "Max DD (%)": round(mdd * 100, 2),
    }


stats_df = pd.DataFrame([
    compute_perf(rp_vw, rf_mon, "Val-Wgt P^(vw)"),
    compute_perf(rp_mv, rf_mon, "Min-Var P_oos^(mv)"),
]).set_index("Portfolio")

print("\n", stats_df.to_string())

# =============================================================================
# 11. VERIFICATION CHECKS
# =============================================================================
print("\n[9] Verification checks ...")

# Check weights sum to 1
for Y, w in mv_w_dict.items():
    wsum = w.sum()
    assert abs(wsum - 1.0) < 1e-6, f"Y={Y}: weights sum to {wsum}, not 1"
print("   ✓ All weight vectors sum to 1")

# Check no negative weights
for Y, w in mv_w_dict.items():
    assert (w >= -1e-8).all(), f"Y={Y}: negative weights found"
print("   ✓ All weights non-negative")

# Check return series length
expected_months = sum(len(months_of(Y + 1)) for Y in range(START_YEAR, END_YEAR + 1))
actual_mv = len(rp_mv.dropna())
actual_vw = len(rp_vw.dropna())
print(f"   Min-Var returns: {actual_mv} months (expected ~{expected_months})")
print(f"   Val-Wgt returns: {actual_vw} months")

# Check no NaN in final return series
assert rp_mv.isna().sum() == 0, "NaN in min-var returns"
assert rp_vw.isna().sum() == 0, "NaN in VW returns"
print("   ✓ No NaN in return series")

# =============================================================================
# 12. TOP HOLDINGS
# =============================================================================
print("\n[10] Top 10 holdings (Min-Var):")

for Y in [2013, 2018, 2024]:
    if Y not in mv_w_dict:
        continue
    w = mv_w_dict[Y].sort_values(ascending=False)
    n_nz = (w > 1e-6).sum()
    print(f"\n   Dec {Y} (non-zero: {n_nz}/{len(w)}):")
    for rk, (isin, wt) in enumerate(w.head(10).items(), 1):
        cty = static.loc[static["ISIN"] == isin, "Country"].values
        cty = cty[0] if len(cty) else ""
        print(f"   {rk:2d}. {isin_name.get(isin, isin):<40s} {cty:>3s}  {wt * 100:6.2f}%")

# =============================================================================
# 13. FIGURES
# =============================================================================
print("\n[11] Generating figures ...")

fig, axes = plt.subplots(2, 2, figsize=(14, 10))
fig.suptitle("SAAM Part I — EUR: Min-Var vs Value-Weighted (2014–2025)",
             fontsize=13, fontweight="bold")
C1, C2 = "steelblue", "darkorange"
fmt = mdates.DateFormatter("%Y")
loc = mdates.YearLocator(2)

# --- Cumulative return ---
ax = axes[0, 0]
cm = (1 + rp_mv.dropna()).cumprod()
cv = (1 + rp_vw.dropna()).cumprod()
ax.plot(cm.index, cm.values, color=C1, lw=1.8, label=r"Min-Var $P_{oos}^{(mv)}$")
ax.plot(cv.index, cv.values, color=C2, lw=1.8, ls="--", label=r"Val-Wgt $P^{(vw)}$")
ax.set_title("Cumulative Return (base=1, Jan 2014)")
ax.set_ylabel("Cumulative return")
ax.legend(fontsize=9)
ax.grid(alpha=0.3)
ax.xaxis.set_major_formatter(fmt)
ax.xaxis.set_major_locator(loc)

# --- Rolling volatility ---
ax = axes[0, 1]
rv_mv = rp_mv.rolling(12).std() * np.sqrt(12) * 100
rv_vw = rp_vw.rolling(12).std() * np.sqrt(12) * 100
ax.plot(rp_mv.index, rv_mv.values, color=C1, lw=1.6, label="Min-Var")
ax.plot(rp_vw.index, rv_vw.values, color=C2, lw=1.6, ls="--", label="Val-Wgt")
ax.set_title("Rolling 12m Annualised Volatility (%)")
ax.set_ylabel("Vol (%)")
ax.legend(fontsize=9)
ax.grid(alpha=0.3)
ax.xaxis.set_major_formatter(fmt)
ax.xaxis.set_major_locator(loc)

# --- Drawdown ---
ax = axes[1, 0]


def drawdown_series(rp):
    c = (1 + rp.dropna()).cumprod()
    return (c - c.cummax()) / c.cummax() * 100


dd_mv = drawdown_series(rp_mv)
dd_vw = drawdown_series(rp_vw)
ax.fill_between(dd_mv.index, dd_mv.values, 0, alpha=0.45, color=C1, label="Min-Var")
ax.fill_between(dd_vw.index, dd_vw.values, 0, alpha=0.30, color=C2, label="Val-Wgt")
ax.set_title("Drawdown from Peak (%)")
ax.set_ylabel("Drawdown (%)")
ax.legend(fontsize=9)
ax.grid(alpha=0.3)
ax.xaxis.set_major_formatter(fmt)
ax.xaxis.set_major_locator(loc)

# --- Universe size ---
ax = axes[1, 1]
yrs = sorted(universe.keys())
ax.bar(yrs, [len(universe[y]) for y in yrs], color=C1, alpha=0.8,
       edgecolor="white", width=0.6)
ax.set_title("EUR Investment Set Size by Year")
ax.set_ylabel("Eligible firms")
ax.set_xticks(yrs)
ax.set_xticklabels(yrs, rotation=45)
ax.grid(axis="y", alpha=0.3)
for y in yrs:
    ax.text(y, len(universe[y]) + 5, str(len(universe[y])),
            ha="center", fontsize=8)

plt.tight_layout()
for ext in ("pdf", "png"):
    plt.savefig(f"{OUT}SAAM_Part1_EUR_figures.{ext}", dpi=150, bbox_inches="tight")
plt.close()
print("   Figures saved.")

# =============================================================================
# 14. EXCEL EXPORT — Official Template Format
# =============================================================================
print("\n[12] Exporting Excel (official template format) ...")
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XlImage

# --- Helper: compute stats as decimals (not percentages) ---
def template_stats(rp, rf_s):
    """Return dict of stats as decimals for the template."""
    rp = rp.dropna()
    rf = rf_s.reindex(rp.index).ffill().fillna(0)
    T = len(rp)
    mu_arith = 12 * rp.mean()                          # slide 12
    mu_geom = (1 + rp).prod() ** (12 / T) - 1          # slide 13
    sig_ann = rp.std() * np.sqrt(12)
    rf_ann = 12 * rf.mean()
    SR = (mu_arith - rf_ann) / sig_ann
    return {
        "ann_avg_ret": mu_arith,      # Row 3: Annualized average return
        "ann_vol": sig_ann,            # Row 4: Annualized volatility
        "ann_cum_ret": mu_geom,        # Row 5: Annualized cumulative return
        "sharpe": SR,                  # Row 6: Sharp ratio
        "min_mo": rp.min(),            # Row 7: Minimum monthly return
        "max_mo": rp.max(),            # Row 8: Maximum monthly return
    }

vw_stats = template_stats(rp_vw, rf_mon)
mv_stats = template_stats(rp_mv, rf_mon)

# --- Fill the template ---
template_path = TEMPLATE_PATH
xlsx_out = f"{OUT}SAAM_Part1_EUR_template.xlsx"

wb = load_workbook(template_path)
ws = wb["Sheet1"]

# Left section: summary statistics (B=VW col 2, C=MV col 3)
stat_keys = ["ann_avg_ret", "ann_vol", "ann_cum_ret", "sharpe", "min_mo", "max_mo"]
for i, key in enumerate(stat_keys):
    ws.cell(row=3 + i, column=2, value=round(vw_stats[key], 8))
    ws.cell(row=3 + i, column=3, value=round(mv_stats[key], 8))

# Row 9: insert cumulative return plot
cum_plot_path = f"{OUT}SAAM_Part1_EUR_cumulative.png"
fig_cum, ax_cum = plt.subplots(figsize=(7, 4))
cm = (1 + rp_mv.dropna()).cumprod()
cv = (1 + rp_vw.dropna()).cumprod()
ax_cum.plot(cm.index, cm.values, color="steelblue", lw=1.8,
            label=r"Min-Var $P_{oos}^{(mv)}$")
ax_cum.plot(cv.index, cv.values, color="darkorange", lw=1.8, ls="--",
            label=r"Val-Wgt $P^{(vw)}$")
ax_cum.set_title("Cumulative Return (base=1, Jan 2014)")
ax_cum.set_ylabel("Cumulative return")
ax_cum.legend(fontsize=9)
ax_cum.grid(alpha=0.3)
ax_cum.xaxis.set_major_formatter(mdates.DateFormatter("%Y"))
ax_cum.xaxis.set_major_locator(mdates.YearLocator(2))
fig_cum.tight_layout()
fig_cum.savefig(cum_plot_path, dpi=150, bbox_inches="tight")
plt.close(fig_cum)

img = XlImage(cum_plot_path)
img.width = 500
img.height = 280
ws.add_image(img, "B9")

# Right section: monthly returns (E=dates col 5, F=VW col 6, G=MV col 7)
# Template already has dates in column E, rows 3 to 146 (144 months)
# Datastream dates are last *business* day, not calendar month-end,
# so we match on (year, month) instead of exact timestamp.
vw_by_ym = {(d.year, d.month): v for d, v in rp_vw.items()}
mv_by_ym = {(d.year, d.month): v for d, v in rp_mv.items()}

for row_idx in range(3, 3 + 144):
    date_cell = ws.cell(row=row_idx, column=5).value
    if date_cell is None:
        continue
    dt = pd.Timestamp(date_cell)
    ym = (dt.year, dt.month)
    vw_val = vw_by_ym.get(ym, np.nan)
    mv_val = mv_by_ym.get(ym, np.nan)
    ws.cell(row=row_idx, column=6, value=round(float(vw_val), 8)
            if not np.isnan(vw_val) else None)
    ws.cell(row=row_idx, column=7, value=round(float(mv_val), 8)
            if not np.isnan(mv_val) else None)

wb.save(xlsx_out)
print(f"   Template saved: {xlsx_out}")

# --- Also save the extended results workbook (for your own reference) ---
xlsx_ext = f"{OUT}SAAM_Part1_EUR_results.xlsx"
with pd.ExcelWriter(xlsx_ext, engine="openpyxl") as writer:
    stats_df.to_excel(writer, sheet_name="Summary_Stats")

    ro = pd.DataFrame({"Min-Var": rp_mv, "Value-Weighted": rp_vw})
    ro.index = ro.index.strftime("%Y-%m")
    ro.to_excel(writer, sheet_name="Monthly_Returns")

    wd = pd.DataFrame(mv_w_dict).T.fillna(0)
    wd.index.name = "Year"
    wd.rename(columns=isin_name, inplace=True)
    wd.to_excel(writer, sheet_name="MV_Weights")

    rows = []
    for Y in sorted(mv_w_dict):
        for rk, (i, wt) in enumerate(
                mv_w_dict[Y].sort_values(ascending=False).head(10).items(), 1
        ):
            cty = static.loc[static["ISIN"] == i, "Country"].values
            rows.append({
                "Year": Y, "Rank": rk, "ISIN": i,
                "Name": isin_name.get(i, i),
                "Country": cty[0] if len(cty) else "",
                "Weight (%)": round(wt * 100, 3),
            })
    pd.DataFrame(rows).to_excel(writer, sheet_name="Top10_Holdings", index=False)

print(f"   Extended results saved: {xlsx_ext}")
print("\n" + "=" * 65)
print("DONE — outputs in", OUT)